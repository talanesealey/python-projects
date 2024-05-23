#Importing of required modules
from openpyxl import load_workbook
import openpyxl
import matplotlib.pyplot as plt

#Note to user within program that explains the usefulness and context of the program.
print("Hi! This is a sample program that examines data for a sample company that I have created. This company has surveyed their workers,")
print("examined their productivity levels, and also surveyed their customers to create information/data on their employees.")
print("This is done to best find out how their company is doing economically. This program is meant to allow the user")
print("to examine and assess this data provided by the company in an easy way for accessibility and usefulness,\nand for hiring and improvement strategies of the company. \n")
print("For this program, you will need to install the following modules: \n openpyxl \n matplotlib.pyplot \n")
print("Thank you! Here is the main menu. \n")

#Defining lists, and opening workbook with data to be read using openpyxl.
wb = load_workbook("SampleDataProject.xlsx")
ws = wb['employee']

nameList = ws['A']
custSatisfaction=ws['B']
workerThoughts=ws['C']
workTurnaround=ws['D']
namesList = []
workerThoughtsList= []
custSatisfactionList= []
workTurnaroundList= []

#Reading excel file columns into parallel lists, i.e. lists with the same length and corresponding information at each index value.
for name in range(len(nameList)):
    if nameList[name].value!="Name":
        namesList.append(nameList[name].value)
for thought in range(len(workerThoughts)):
    if workerThoughts[thought].value!="Workers Thoughts":
        workerThoughtsList.append(workerThoughts[thought].value)
for satisfaction in range(len(custSatisfaction)):
    if custSatisfaction[satisfaction].value != "Customer Satisfaction":
        custSatisfactionList.append(custSatisfaction[satisfaction].value)
for turnaround in range(len(workTurnaround)):
    if workTurnaround[turnaround].value!="Work Turnaround (in weeks)":
        workTurnaroundList.append(int(workTurnaround[turnaround].value))

#CODING FOR EACH MENU OPTION BELOW: Option 1, 2, and 3 and 4 are defined in functions as option1(), option2(), option3() and option4().
#                                    The actual menu input is created at the end of the program.


#Menu Option 1: This option asks the user for the desired category they wish to search through. It is exhaustive searching that
#               provides category descriptions depending on chosen category, and seperates each category into subcategory to isolate the worker in each subcategory. It then returns to the main menu.
def option1():
    greatEmployees=[]
    goodEmployees=[]
    badEmployees=[]
    likeable=[]
    dislikeable=[]
    bestEmployee=[]
    worstEmployee=[]
    print()
    category=input("Enter number of category that you would like: \n 1. Customer Satisfaction \n 2. Worker Thoughts \n 3. Work Turnaround \n 4. Find best worker and worst worker. \n")
    if category=='1':
        choice=False
        print("Description: Customer Satisfaction describes the level of satisfaction that customers have expressed about the worker (through surveying)")
        for idx in range(len(custSatisfactionList)):
                if custSatisfactionList[idx].lower().strip()=="great":
                    greatEmployees.append(namesList[idx])
                if custSatisfactionList[idx].lower().strip()=="good":
                    goodEmployees.append(namesList[idx])
                if custSatisfactionList[idx].lower().strip()=="bad":
                    badEmployees.append(namesList[idx])
        print("Through an assessment of customer satisfaction,")
        print("The great employees are:")
        for idx in range(len(greatEmployees)):
            print(greatEmployees[idx], end=', ')
        print("\n")
        print("The good employees are:")
        for idx in range(len(goodEmployees)):
            print(goodEmployees[idx], end=', ')
        print("\n")
        print("The bad employees are:")
        for idx in range(len(badEmployees)):
            print(badEmployees[idx], end=', ')
    elif category=='2':
        print("Description: Worker Thoughts describes the likeability of the worker from the perspective of their colleagues, or other workers.")
        likeable=[]
        dislikeable=[]
        for idx in range(len(workerThoughtsList)):
            if workerThoughtsList[idx].lower().strip()=="likeable":
                likeable.append(namesList[idx])
            if workerThoughtsList[idx].lower().strip()=="dislikeable":
                dislikeable.append(namesList[idx])
        print("The likeable employees are:")
        for idx in range(len(likeable)):
            print(likeable[idx], end=', ')
        print("\n")
        print("The dislikeable employees are:")
        for idx in range(len(dislikeable)):
            print(dislikeable[idx], end=', ')
    elif category=='3':
       print("Description: Worker Turnaround is how quick the worker is able to turn in their assigned projects or tasks in the workplace (in weeks). A worker that has a turnaround of less than or equal to 2 weeks is considered satisfactory.")
       print("The worker's turnaround ranges from 1 to 4 weeks")
       print("Here is a list of the work turnaround numbers for each worker:")
       for idx in range(len(workTurnaroundList)):
                print(workTurnaroundList[idx], end=', ')
    elif category=='4':
        minWorkTurnAround = 2
        for min in range(len(workTurnaroundList)):
            if  workTurnaroundList[min] <= minWorkTurnAround and workerThoughtsList[min].lower()=='likeable' and custSatisfactionList[min].lower()=='great':
                bestEmployee.append(namesList[min])
        print("The best employee is {}, with a work turnaround of 2, a worker thought rating of likeable and a customer satisfaction rating of great.".format(bestEmployee[0]))
        for max in range(len(workTurnaroundList)):
            if workTurnaroundList[max] > minWorkTurnAround and workerThoughtsList[max].lower()=='dislikeable' and custSatisfactionList[max].lower()=='bad':
                worstEmployee.append(namesList[max])
        print("The worst employees are {}, with a work turnaround greater than 2, a worker thought rating of dislikeable and a customer satisfaction rating of bad.".format(worstEmployee))
    print()
    print("\nGoing back to main menu... \n")
    menu()
#Menu Option 2: This option asks the user which worker they would like to examine and then produces their worker statistics including
#                likeability, customer satisfaction, and work turnaround to be seen by the user. It is a match and stop search that assesses
#                wrong input options.
def option2():
    done=True
    name=0
    print("Here is a list of the workers: {}".format(namesList))
    while done==True:
        inputWorker=input("Which worker would you like to examine? ")
        print()
        if inputWorker not in namesList:
            print("No workers with that name exist. Please re-enter.")
        else:
            while name<len(namesList) and done==True:
                if namesList[name].lower()==inputWorker.lower():
                     done=False
                name+=1
            print("The worker's statistics for {} are:".format(namesList[name-1]))
            print()
            print("Worker's Thoughts: {}".format(workerThoughtsList[name-1]))
            print("Worker's Turnaround (in weeks): {}".format(workTurnaroundList[name-1]))
            print("Customer Satisfaction: {}".format(custSatisfactionList[name-1]))
    print("\nGoing back to main menu... \n")
    menu()
#Menu Option 3: This option gives the user the option to plot data. It asks the user which y axis category they would plot and
#               creates a bar chart for them, based on their option. It also gives the user the option to save it to an .xlsx file
#               for reference later. It then returns to the main menu.
def option3():
    greatEmployees=[]
    goodEmployees=[]
    badEmployees=[]
    likeable=[]
    dislikeable=[]
    likeablecount=0
    dislikeablecount=0
    greatEmployeesCount=0
    goodEmployeesCount=0
    badEmployeesCount=0
    inputXAxis=0
    lessthan=0
    greaterthan=0
    print("Categories:\n 1. Worker's Thoughts \n 2. Worker's Turnaround (in weeks) \n 3. Customer's Satisfaction")
    print()
    inputYAxis=int(input("Enter desired y axis from categories that you would like to plot. (Enter category number only) \n "))
    if inputYAxis==1:
         for idx in range(len(workerThoughtsList)):
            if workerThoughtsList[idx].lower().strip()=="likeable":
                likeable.append(namesList[idx])
                likeablecount+=1
            if workerThoughtsList[idx].lower().strip()=="dislikeable":
                dislikeable.append(namesList[idx])
                dislikeablecount+=1
            inputYAxis=[likeablecount, dislikeablecount]
            inputXAxis=['Likeable','Dislikeable']
            xAxisLabel='Level of Likeability'
    elif inputYAxis==2:
        print("\nNOTE: The criteria for a good worker is a worker that submits assignment in less than or equal to 2 weeks. \n ")
        for idx in range(len(workTurnaroundList)):
            if workTurnaroundList[idx]<=2:
                lessthan+=1
            else:
                greaterthan+=1
        inputYAxis=[lessthan,greaterthan]
        inputXAxis=['Less than or equal to 2 weeks', 'Greater than 2 weeks']
        xAxisLabel='Length of Work Turnaround'
    elif inputYAxis==3:
        for idx in range(len(custSatisfactionList)):
            if custSatisfactionList[idx].lower().strip()=="great":
                greatEmployees.append(namesList[idx])
                greatEmployeesCount+=1
            if custSatisfactionList[idx].lower().strip()=="good":
                goodEmployees.append(namesList[idx])
                goodEmployeesCount+=1
            if custSatisfactionList[idx].lower().strip()=="bad":
                badEmployees.append(namesList[idx])
                badEmployeesCount+=1
            inputYAxis=[greatEmployeesCount,goodEmployeesCount,badEmployeesCount]
            inputXAxis=['Great','Good','Bad']
            xAxisLabel='Level of Customer Satisfaction'
    inputPlotTitle=input("Enter desired title for chart: ")
    plt.bar(inputXAxis, inputYAxis, color=['pink', 'orange'])
    plt.xlabel(xAxisLabel)
    plt.ylabel('Number of Workers')
    plt.title(inputPlotTitle)
    plt.savefig('myplot.png', dpi = 150)
    plt.show()
    inputExcel=input("Would you like this graph to be saved to an Excel Worksheet for reference? (Yes or No) ")
    if inputExcel.lower()=="yes":
        workbook = openpyxl.Workbook()
        workbook.save("mypythonplot.xlsx")
        wb = openpyxl.load_workbook("mypythonplot.xlsx")
        ws = wb.active
        img = openpyxl.drawing.image.Image('myplot.png')
        img.anchor = 'A1'
        ws.add_image(img)
        wb.save("mypythonplot.xlsx")
        print("\nYour file was saved as mypythonplot.xlsx in the same path as this python file for reference. Enjoy viewing!")
    print("\nGoing back to main menu... \n")
    menu()
#Menu Option 4: This menu option provides user with background information on the categories assessed in this program.
def option4():
    print("Categories:\n 1. Worker's Thoughts \n 2. Worker's Turnaround (in weeks) \n 3. Customer's Satisfaction")
    print()
    categoryChoice=input("Enter category number above for selected description: ")
    if categoryChoice=='1':
        print("Worker Thoughts describes the likeability of the worker from the perspective of their colleagues, or other workers. Workers are defined as either likeable or dislikeable.")
    elif categoryChoice=='2':
     print("Worker Turnaround is how quick the worker is able to turn in their assigned projects or tasks in the workplace (in weeks). A worker that has a turnaround of less than or equal to 2 weeks is considered satisfactory.")
     print("The worker's turnaround ranges from {} to {} weeks".format(min(workTurnaroundList),max(workTurnaroundList)))
    elif categoryChoice=='3':
        print("Customer Satisfaction describes the level of satisfaction that customers have expressed about the worker (through surveying). It is measured on a scale of Great, Good, and Bad.")
    print("\nGoing back to main menu... \n")
    menu()
#Menu Option 5: Allows user to quit program and exit.
def option5():
    print("\n Quitting program. Thank you!")
    quit()
#Overhaul Menu that allows user to pick through 4 options that all do different coding techniques.
def menu():
    menu=input("MAIN MENU: \n 1. Search through categories: Customer Satisfaction, Worker Thoughts, Work Turnaround, Best & Worst Employees \n 2. Search for a worker and their statistics \n 3. Plot collected worker data \n 4. Description of Categories and other information. \n 5. Exit \n")
    if menu=='1':
        option1()
    elif menu=='2':
        option2()
    elif menu=='3':
        option3()
    elif menu=='4':
        option4()
    elif menu=='5':
        option5()
menu()
